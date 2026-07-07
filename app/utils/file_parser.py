"""
Utility for parsing uploaded Excel (.xlsx) and CSV files into row dictionaries.

Provides a single entry point that detects the file type, reads its contents,
and returns a list of dictionaries keyed by the file's header row. The parser
is table-agnostic — it simply converts a spreadsheet into raw string rows,
consistent with Bronze-layer ingestion. Validation against a specific table's
schema happens later, in the service layer.
"""

import csv
import io

from openpyxl import load_workbook

from app.utils.logger import get_logger

logger = get_logger(__name__)


class FileParseError(Exception):
    """Raised when an uploaded file cannot be parsed."""


def parse_csv(content: bytes) -> list[dict[str, str]]:
    """
    Parse CSV file bytes into a list of row dictionaries.

    Args:
        content: The raw bytes of the uploaded CSV file.

    Returns:
        list[dict[str, str]]: One dict per data row, keyed by header names.

    Raises:
        FileParseError: If the file is empty or cannot be decoded.
    """
    try:
        text = content.decode("utf-8-sig")  # utf-8-sig strips Excel's BOM
    except UnicodeDecodeError as exc:
        raise FileParseError("File is not valid UTF-8 text.") from exc

    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise FileParseError("CSV file has no header row.")

    rows = [dict(row) for row in reader]
    if not rows:
        raise FileParseError("CSV file contains no data rows.")
    return rows


def parse_xlsx(content: bytes) -> list[dict[str, str]]:
    """
    Parse Excel (.xlsx) file bytes into a list of row dictionaries.

    Reads the first worksheet. The first row is treated as the header;
    each subsequent row becomes a dictionary keyed by those headers.

    Args:
        content: The raw bytes of the uploaded .xlsx file.

    Returns:
        list[dict[str, str]]: One dict per data row, keyed by header names.

    Raises:
        FileParseError: If the file is empty or malformed.
    """
    try:
        workbook = load_workbook(
            filename=io.BytesIO(content), read_only=True, data_only=True
        )
    except Exception as exc:
        raise FileParseError("Could not read the Excel file.") from exc

    worksheet = workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration:
        raise FileParseError("Excel file is empty.") from None

    if header is None or all(cell is None for cell in header):
        raise FileParseError("Excel file has no header row.")

    headers = [str(cell).strip() if cell is not None else "" for cell in header]

    rows: list[dict[str, str]] = []
    for row in rows_iter:
        # Skip fully-empty rows.
        if row is None or all(cell is None for cell in row):
            continue
        # Convert every cell to string (Bronze layer = raw strings).
        row_dict = {
            headers[i]: ("" if row[i] is None else str(row[i]))
            for i in range(len(headers))
        }
        rows.append(row_dict)

    if not rows:
        raise FileParseError("Excel file contains no data rows.")
    return rows


def parse_upload(filename: str, content: bytes) -> list[dict[str, str]]:
    """
    Detect the file type by extension and parse accordingly.

    Args:
        filename: The uploaded file's name (used to detect .csv vs .xlsx).
        content: The raw bytes of the uploaded file.

    Returns:
        list[dict[str, str]]: Parsed rows as dictionaries.

    Raises:
        FileParseError: If the extension is unsupported or parsing fails.
    """
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        logger.info("Parsing uploaded CSV file: %s", filename)
        return parse_csv(content)
    if lower_name.endswith(".xlsx"):
        logger.info("Parsing uploaded Excel file: %s", filename)
        return parse_xlsx(content)

    raise FileParseError(
        "Unsupported file type. Please upload a .csv or .xlsx file."
    )